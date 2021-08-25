from openpyxl import load_workbook

def read_cases(excel_path, excel_sheet):
    wb = load_workbook(excel_path)
    sheet = wb[excel_sheet]
    cases = []
    for i in range(2, sheet.max_row+1):
         case_id = sheet.cell(row=i, column=1).value
         url = sheet.cell(row=i, column=5).value
         data = sheet.cell(row=i, column=6).value
         expected = sheet.cell(row=i, column=7).value
         cases.append(dict(case_id=case_id,url=url,data=data,expected=expected))
    return cases

def write_cases(excel_path, excel_sheet, results):
    wb = load_workbook(excel_path)
    sheet = wb[excel_sheet]
    for i in range(2, sheet.max_row+1):
        sheet.cell(row=i, column=8).value = results[i-2]
    wb.save(excel_path)




